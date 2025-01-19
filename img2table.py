import os
import cv2
import openpyxl
from paddleocr import PPStructure,draw_structure_result,save_structure_res
import shutil
import regex

TEMP_SAVES = "/tmp_ocr/"

def get_save_folder():
    return os.getcwd() + TEMP_SAVES

def get_front_file(dire):
    front = dire.rfind('\\')
    back = dire.rfind('/')
    if front == -1 and back == -1:
        return dire
    if front > back:
        return dire[dire.rfind('\\'):]
    else:
        return dire[dire.rfind('/'):]
    
def find_spreadsheet_file(innermost_dir):
    for file in os.listdir(innermost_dir):
        print(file)
        if file.endswith(".xlsx"):
            return file
    return None


def read_image(img_path):

    def _contains_english(i):
        return regex.search(r'[A-Za-z]', i)
    
    def _contains_chinese(i):
        return regex.search(r'\p{Han}+', i)
    

    table_engine = PPStructure(show_log=True, image_orientation=False)

    tempsave_folder = './output'
    image1 = cv2.imread(img_path)
    result = table_engine(image1)
    t = save_structure_res(result, get_save_folder(),os.path.basename(img_path).split('.')[0])

    print(os.path.basename(img_path).split('.')[0])
    #p = get_front_file(img_path)

    pathtofolder = get_save_folder() + os.path.basename(img_path).split('.')[0]
    
    print(pathtofolder)

    sheet = openpyxl.load_workbook(pathtofolder + "/" + find_spreadsheet_file(pathtofolder)).active

    raw_out = []
    for row in range(0, sheet.max_row):
        tr = []
        for col in sheet.iter_cols(1, sheet.max_column):
           tr.append((col[row].value))
        #print(tr)
        raw_out.append(tr)


    shutil.rmtree(pathtofolder)

    out = []
    for row in raw_out:
        t = []
        for cell in row:
            if not bool(cell):
                continue
            
            chinese = _contains_chinese(cell)
            english = _contains_english(cell)
            

            if chinese and english:
                cn_part = ""
                en_part = ""

                parts = cell.split()

                for part in parts:
                    if _contains_chinese(part):
                        cn_part = part
                    else:
                        en_part = part
                t.append(cn_part)
                t.append(en_part)
                continue
            else:
                t.append(cell)


        out.append(t)
    
    return out


    
    





    
    
    
    


if __name__ == "__main__":
    for row in read_image("test3.png"):
        print(row)


# for line in result:
#     line.pop('img')
#     print(line)

# from PIL import Image

# font_path = 'doc/fonts/simfang.ttf' 
# image = Image.open(img_path).convert('RGB')
# im_show = draw_structure_result(image, result,font_path=font_path)
# im_show = Image.fromarray(im_show)
# im_show.save('result.jpg')
